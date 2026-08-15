import openpyxl


path = r"D:/Capstone_FoodResQ_Web/outputs/report5-foodresq/FoodResQ_Report5_Test_Report_v11.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

ws = wb["Test Cases"]
print("test_cases_hyperlink_formulas=")
for row in range(9, 15):
    print(row, ws[f"D{row}"].value)

for sheet_name in ["Login", "Campaign Kitchen Ops", "Delivery Shipper"]:
    ws = wb[sheet_name]
    sections = []
    tc_count = 0
    for row in range(11, 80):
        value = ws[f"A{row}"].value
        if isinstance(value, str) and value:
            if value.startswith("TC"):
                tc_count += 1
            else:
                sections.append((row, value))
    print(sheet_name, "sections=", sections, "tc_count=", tc_count, "B4=", ws["B4"].value)
