import openpyxl


path = r"D:/Capstone_FoodResQ_Web/outputs/report5-foodresq/FoodResQ_Report5_Test_Report_v11.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("sheets=", wb.sheetnames)

exclude = {"Cover", "Test Cases", "Test Statistics"}
module_counts = []
total = 0
for sheet_name in wb.sheetnames:
    if sheet_name in exclude:
        continue
    ws = wb[sheet_name]
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=160, max_col=1, values_only=True):
        value = row[0]
        if isinstance(value, str) and value.strip().upper().startswith("TC"):
            count += 1
    if count:
        module_counts.append((sheet_name, count))
        total += count

print("module_counts=", module_counts)
print("total=", total)

if "Test Statistics" in wb.sheetnames:
    print("test_statistics_rows=")
    ws = wb["Test Statistics"]
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=10, values_only=True):
        vals = [value for value in row if value is not None]
        if vals:
            print(vals)
